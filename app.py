# -*- coding: utf-8 -*-
import os, io, re, wave, subprocess, shutil
import numpy as np
import requests
from flask import Flask, request, jsonify, send_file, send_from_directory

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
# Keys come from environment variables (set them on Render, or in a local .env / shell).
# They are intentionally NOT hardcoded — GitHub blocks pushing recognized secrets, and
# committing keys to a repo exposes them publicly.
ELEVEN_KEY = os.environ.get("ELEVEN_API_KEY", "")
AZURE_KEY = os.environ.get("AZURE_TTS_KEY", "")
AZURE_REGION = os.environ.get("AZURE_TTS_REGION", "eastus")
EL_BASE = "https://api.elevenlabs.io"
SAMPLE_RATE = 44100

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
GEN_DIR = os.path.join(BASE_DIR, "generated")
ASSET_DIR = os.path.join(BASE_DIR, "assets")
os.makedirs(GEN_DIR, exist_ok=True)

# ffmpeg: use system binary if present, else the pip-bundled one (imageio-ffmpeg)
def _resolve_ffmpeg():
    exe = shutil.which("ffmpeg")
    if exe:
        return exe
    try:
        import imageio_ffmpeg
        return imageio_ffmpeg.get_ffmpeg_exe()
    except Exception:
        return "ffmpeg"
FFMPEG = _resolve_ffmpeg()

app = Flask(__name__, static_folder=None)


# ---------------------------------------------------------------------------
# WAV helpers (internal working format: 16-bit mono PCM @ 44100)
# ---------------------------------------------------------------------------
def pcm_to_wav(pcm_bytes, sample_rate=SAMPLE_RATE):
    buf = io.BytesIO()
    with wave.open(buf, "wb") as w:
        w.setnchannels(1); w.setsampwidth(2); w.setframerate(sample_rate)
        w.writeframes(pcm_bytes)
    return buf.getvalue()


def read_wav_int16(path):
    with wave.open(path, "rb") as w:
        n, ch, sr = w.getnframes(), w.getnchannels(), w.getframerate()
        raw = w.readframes(n)
    data = np.frombuffer(raw, dtype=np.int16).astype(np.float32)
    if ch == 2:
        data = data.reshape(-1, 2).mean(axis=1)
    if sr != SAMPLE_RATE and len(data) > 1:
        data = np.interp(np.linspace(0, 1, int(round(len(data) * SAMPLE_RATE / sr))),
                         np.linspace(0, 1, len(data)), data)
    return data


def write_wav_int16(path, samples, sample_rate=SAMPLE_RATE):
    clipped = np.clip(samples, -32768, 32767).astype(np.int16)
    with wave.open(path, "wb") as w:
        w.setnchannels(1); w.setsampwidth(2); w.setframerate(sample_rate)
        w.writeframes(clipped.tobytes())


def resample_mono_44100(path):
    with wave.open(path, "rb") as w:
        n, ch, sr, sw = w.getnframes(), w.getnchannels(), w.getframerate(), w.getsampwidth()
        raw = w.readframes(n)
    if sw == 2:
        data = np.frombuffer(raw, dtype=np.int16).astype(np.float32)
    elif sw == 1:
        data = (np.frombuffer(raw, dtype=np.uint8).astype(np.float32) - 128) * 256
    elif sw == 4:
        data = (np.frombuffer(raw, dtype=np.int32).astype(np.float32)) / 65536.0
    else:
        raise ValueError(f"Unsupported width {sw*8}-bit")
    if ch > 1:
        data = data.reshape(-1, ch).mean(axis=1)
    if sr != SAMPLE_RATE and len(data) > 1:
        data = np.interp(np.linspace(0, 1, int(round(len(data) * SAMPLE_RATE / sr))),
                         np.linspace(0, 1, len(data)), data)
    return data


def wav_duration_seconds(path):
    with wave.open(path, "rb") as w:
        return w.getnframes() / float(w.getframerate())


def safe_name(name):
    name = re.sub(r'[<>:"/\\|?*\n\r]', "", str(name).strip())
    name = re.sub(r"\s+", "_", name)
    return name or "output"


def clean_sid(sid):
    sid = re.sub(r"[^A-Za-z0-9_-]", "", str(sid or ""))[:40]
    return sid or "anon"


def xml_escape(t):
    return (str(t).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            .replace('"', "&quot;").replace("'", "&apos;"))


# ---------------------------------------------------------------------------
# TTS engines  -> return (float samples @44100 mono, error)
# ---------------------------------------------------------------------------
def tts_elevenlabs(d, sid):
    voice_id = (d.get("voice_id") or "").strip()
    text = (d.get("text") or "").strip()
    if not voice_id or not text:
        return None, "Voice ID and text are required."
    key = (d.get("api_key") or "").strip() or ELEVEN_KEY
    model = d.get("model_id", "eleven_flash_v2_5")
    vs = {"stability": float(d.get("stability", 1.0)),
          "similarity_boost": float(d.get("similarity_boost", 1.0))}
    if model != "eleven_v3":
        vs["style"] = float(d.get("style", 0.0))
        vs["speed"] = float(d.get("speed", 1.0))
    url = f"{EL_BASE}/v1/text-to-speech/{voice_id}?output_format=pcm_44100"
    payload = {"text": text, "model_id": model, "voice_settings": vs,
               "apply_text_normalization": d.get("apply_text_normalization", "off")}
    headers = {"xi-api-key": key, "Content-Type": "application/json"}
    try:
        r = requests.post(url, headers=headers, json=payload, timeout=180)
    except Exception as e:
        return None, f"Request failed: {e}"
    if r.status_code != 200:
        return None, f"ElevenLabs error {r.status_code}: {r.text[:300]}"
    tmp = os.path.join(GEN_DIR, f"{sid}_el_tmp.wav")
    with open(tmp, "wb") as f:
        f.write(pcm_to_wav(r.content))
    return resample_mono_44100(tmp), None


AZURE_STYLE_NONE = "(none)"

def tts_azure(d, sid):
    text = (d.get("text") or "").strip()
    voice = (d.get("azure_voice") or "en-US-JennyNeural").strip()
    if not text:
        return None, "Text is required."
    key = (d.get("azure_key") or "").strip() or AZURE_KEY
    region = (d.get("azure_region") or AZURE_REGION).strip()
    locale = (d.get("azure_locale") or "en-US").strip()
    style = (d.get("azure_style") or "").strip()
    style_degree = d.get("azure_style_degree", "1")
    rate = d.get("azure_rate", "0")

    inner = xml_escape(text)
    try:
        rate_i = int(float(rate))
    except Exception:
        rate_i = 0
    if rate_i != 0:
        sign = "+" if rate_i > 0 else ""
        inner = f'<prosody rate="{sign}{rate_i}%">{inner}</prosody>'
    if style and style != AZURE_STYLE_NONE:
        inner = f'<mstts:express-as style="{xml_escape(style)}" styledegree="{xml_escape(style_degree)}">{inner}</mstts:express-as>'
    ssml = (f'<speak version="1.0" xmlns="http://www.w3.org/2001/10/synthesis" '
            f'xmlns:mstts="https://www.w3.org/2001/mstts" xml:lang="{xml_escape(locale)}">'
            f'<voice name="{xml_escape(voice)}">{inner}</voice></speak>')

    url = f"https://{region}.tts.speech.microsoft.com/cognitiveservices/v1"
    headers = {"Content-Type": "application/ssml+xml", "Ocp-Apim-Subscription-Key": key,
               "X-Microsoft-OutputFormat": "riff-24khz-16bit-mono-pcm", "User-Agent": "TTSStudio"}
    try:
        r = requests.post(url, headers=headers, data=ssml.encode("utf-8"), timeout=120)
    except Exception as e:
        return None, f"Request failed: {e}"
    if r.status_code != 200:
        return None, f"Azure error {r.status_code}: {r.text[:300]}"
    tmp = os.path.join(GEN_DIR, f"{sid}_az_tmp.wav")
    with open(tmp, "wb") as f:
        f.write(r.content)
    return resample_mono_44100(tmp), None


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------
@app.route("/")
def index():
    return send_file(os.path.join(BASE_DIR, "index.html"))


@app.route("/health")
def health():
    return jsonify(ok=True, ffmpeg=FFMPEG)


@app.route("/audio/<path:fname>")
def audio(fname):
    return send_from_directory(GEN_DIR, fname)


@app.route("/api/tts", methods=["POST"])
def api_tts():
    d = request.get_json(force=True)
    sid = clean_sid(d.get("sid"))
    engine = (d.get("engine") or "elevenlabs").lower()
    samples, err = tts_azure(d, sid) if engine == "azure" else tts_elevenlabs(d, sid)
    if err:
        return jsonify(error=err), 502
    fname = f"{sid}_speech.wav"
    write_wav_int16(os.path.join(GEN_DIR, fname), samples)
    return jsonify(file=fname, duration=round(len(samples) / SAMPLE_RATE, 2))


@app.route("/api/use-default-music", methods=["POST"])
def use_default_music():
    d = request.get_json(force=True)
    sid = clean_sid(d.get("sid"))
    src = os.path.join(ASSET_DIR, "default_music.wav")
    if not os.path.exists(src):
        return jsonify(error="Default music not found on server."), 404
    samples = resample_mono_44100(src)
    write_wav_int16(os.path.join(GEN_DIR, f"{sid}_music_orig.wav"), samples)
    write_wav_int16(os.path.join(GEN_DIR, f"{sid}_music.wav"), samples)
    return jsonify(file=f"{sid}_music_orig.wav", duration=round(len(samples) / SAMPLE_RATE, 2))


@app.route("/api/upload-music", methods=["POST"])
def upload_music():
    if "file" not in request.files:
        return jsonify(error="No file uploaded."), 400
    sid = clean_sid(request.form.get("sid"))
    tmp = os.path.join(GEN_DIR, f"{sid}_upload_music.wav")
    request.files["file"].save(tmp)
    try:
        samples = resample_mono_44100(tmp)
    except Exception as e:
        return jsonify(error=f"Could not read audio (WAV only): {e}"), 400
    write_wav_int16(os.path.join(GEN_DIR, f"{sid}_music_orig.wav"), samples)
    write_wav_int16(os.path.join(GEN_DIR, f"{sid}_music.wav"), samples)
    return jsonify(file=f"{sid}_music_orig.wav", duration=round(len(samples) / SAMPLE_RATE, 2))


@app.route("/api/trim-music", methods=["POST"])
def trim_music():
    d = request.get_json(force=True)
    sid = clean_sid(d.get("sid"))
    orig = os.path.join(GEN_DIR, f"{sid}_music_orig.wav")
    if not os.path.exists(orig):
        return jsonify(error="Load or upload music first."), 400
    samples = read_wav_int16(orig)
    total = len(samples) / float(SAMPLE_RATE)
    start = max(0.0, min(float(d.get("start", 0)), total))
    end = max(start, min(float(d.get("end", total)), total))
    a, b = int(round(start * SAMPLE_RATE)), int(round(end * SAMPLE_RATE))
    if b - a < int(0.05 * SAMPLE_RATE):
        return jsonify(error="Selection is too short."), 400
    write_wav_int16(os.path.join(GEN_DIR, f"{sid}_music.wav"), samples[a:b])
    return jsonify(file=f"{sid}_music.wav", duration=round((b - a) / float(SAMPLE_RATE), 2))


FORMATS = {
    "pcm_44100": ["-ar", "44100", "-ac", "1", "-c:a", "pcm_s16le"],
    "pcm_16000": ["-ar", "16000", "-ac", "1", "-c:a", "pcm_s16le"],
    "pcm_8000":  ["-ar", "8000",  "-ac", "1", "-c:a", "pcm_s16le"],
    "alaw_8000": ["-ar", "8000",  "-ac", "1", "-c:a", "pcm_alaw"],
    "ulaw_8000": ["-ar", "8000",  "-ac", "1", "-c:a", "pcm_mulaw"],
}


def synth(d, sid):
    """Dispatch to the chosen TTS engine -> (float samples @44100 mono, error)."""
    engine = (d.get("engine") or "elevenlabs").lower()
    return tts_azure(d, sid) if engine == "azure" else tts_elevenlabs(d, sid)


def mix_and_encode(sid, speech, use_music, music_vol, fmt, out_base):
    """Mix optional music bed under speech and encode to the target format.
    Returns (out_name, download_name, duration, error)."""
    if fmt not in FORMATS:
        return None, None, None, f"Unknown format {fmt}."
    out_base = safe_name(out_base)
    if out_base.lower().endswith(".wav"):
        out_base = out_base[:-4]
    out_name = f"{sid}__{out_base}.wav"

    N = len(speech)
    mix = speech.copy()
    music_path = os.path.join(GEN_DIR, f"{sid}_music.wav")
    if use_music and os.path.exists(music_path):
        music = read_wav_int16(music_path)
        if len(music) > 0:
            bed = np.tile(music, int(np.ceil(N / len(music))))[:N] if len(music) < N else music[:N].copy()
            fade = int(0.4 * SAMPLE_RATE)
            if N > 2 * fade and fade > 0:
                bed[:fade] *= np.linspace(0, 1, fade)
                bed[-fade:] *= np.linspace(1, 0, fade)
            mix = speech + music_vol * bed
            peak = float(np.max(np.abs(mix))) if N else 0
            if peak > 32767:
                mix *= 32767.0 / peak

    tmp = os.path.join(GEN_DIR, f"{sid}_mix_tmp.wav")
    write_wav_int16(tmp, mix)
    out_path = os.path.join(GEN_DIR, out_name)
    cmd = [FFMPEG, "-y", "-i", tmp] + FORMATS[fmt] + ["-fflags", "+bitexact", out_path]
    p = subprocess.run(cmd, capture_output=True, text=True)
    if p.returncode != 0:
        return None, None, None, f"ffmpeg failed: {p.stderr[-300:]}"
    return out_name, f"{out_base}.wav", round(N / float(SAMPLE_RATE), 2), None


@app.route("/api/export", methods=["POST"])
def api_export():
    d = request.get_json(force=True)
    sid = clean_sid(d.get("sid"))
    sp_path = os.path.join(GEN_DIR, f"{sid}_speech.wav")
    if not os.path.exists(sp_path):
        return jsonify(error="Generate speech first."), 400
    speech = read_wav_int16(sp_path)
    out_name, dl, dur, err = mix_and_encode(
        sid, speech, bool(d.get("use_music")),
        float(d.get("music_volume", 15)) / 100.0,
        d.get("format", "pcm_8000"), d.get("output_name") or "output")
    if err:
        return jsonify(error=err), 500
    return jsonify(file=out_name, download_name=dl, duration=dur, format=d.get("format", "pcm_8000"))


@app.route("/api/render", methods=["POST"])
def api_render():
    """Generate speech for arbitrary text + settings and export in one call.
    Used by the Bulk tab and per-row regeneration."""
    d = request.get_json(force=True)
    sid = clean_sid(d.get("sid"))
    samples, err = synth(d, sid)
    if err:
        return jsonify(error=err), 502
    out_name, dl, dur, err = mix_and_encode(
        sid, samples, bool(d.get("use_music")),
        float(d.get("music_volume", 15)) / 100.0,
        d.get("format", "pcm_8000"), d.get("output_name") or "output")
    if err:
        return jsonify(error=err), 500
    return jsonify(file=out_name, download_name=dl, duration=dur, format=d.get("format", "pcm_8000"))


@app.route("/api/parse-sheet", methods=["POST"])
def parse_sheet():
    if "file" not in request.files:
        return jsonify(error="No file uploaded."), 400
    f = request.files["file"]
    fname = (f.filename or "").lower()
    data = f.read()
    MAX_ROWS = 2000
    sheets = []
    try:
        if fname.endswith(".csv"):
            import csv, io as _io
            txt = data.decode("utf-8-sig", errors="replace")
            reader = list(csv.reader(_io.StringIO(txt)))
            reader = [r for r in reader if any((c or "").strip() for c in r)]
            headers = [str(c) for c in reader[0]] if reader else []
            rows = [[("" if c is None else str(c)) for c in r] for r in reader[1:MAX_ROWS + 1]]
            sheets.append({"name": "CSV", "columns": headers, "rows": rows})
        else:
            import openpyxl, io as _io
            wb = openpyxl.load_workbook(_io.BytesIO(data), read_only=True, data_only=True)
            for sname in wb.sheetnames:
                ws = wb[sname]
                allrows = list(ws.iter_rows(values_only=True))
                if not allrows:
                    sheets.append({"name": sname, "columns": [], "rows": []}); continue
                headers = [("" if c is None else str(c)) for c in allrows[0]]
                rows = [[("" if c is None else str(c)) for c in r] for r in allrows[1:MAX_ROWS + 1]]
                sheets.append({"name": sname, "columns": headers, "rows": rows})
    except Exception as e:
        return jsonify(error=f"Could not read sheet: {e}"), 400
    return jsonify(sheets=sheets)


@app.route("/api/bulk-zip")
def bulk_zip():
    sid = clean_sid(request.args.get("sid"))
    import zipfile, io as _io
    prefix = f"{sid}__"
    buf = _io.BytesIO()
    count = 0
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for fn in sorted(os.listdir(GEN_DIR)):
            if fn.startswith(prefix) and fn.endswith(".wav"):
                z.write(os.path.join(GEN_DIR, fn), arcname=fn[len(prefix):])
                count += 1
    if count == 0:
        return jsonify(error="No generated files yet."), 400
    buf.seek(0)
    return send_file(buf, mimetype="application/zip", as_attachment=True, download_name="tts_bulk.zip")


@app.route("/api/download/<path:fname>")
def download(fname):
    dn = request.args.get("name", fname)
    return send_from_directory(GEN_DIR, fname, as_attachment=True, download_name=dn)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5001))
    print(f"TTS Studio running at http://127.0.0.1:{port}  (ffmpeg: {FFMPEG})")
    app.run(host="0.0.0.0", port=port, debug=False)
